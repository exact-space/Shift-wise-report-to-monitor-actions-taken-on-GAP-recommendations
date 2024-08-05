from apscheduler.schedulers.background import BackgroundScheduler
import threading
import requests,json
import pandas as pd 
from fuzzywuzzy import process
import re
import pprint
import datetime
from datetime import datetime, timedelta
import datetime as datetime
from pytz import timezone
import pytz
import time
from fuzzywuzzy import process
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import platform
import os
import messaging as mg
version = platform.python_version().split(".")[0]
if version == "3":
  import app_config.app_config as cfg
elif version == "2":
    import app_config as cfg
config = cfg.getconfig()

# unitId = os.environ.get("UNIT_ID")
# unitId = "60ae9143e284d016d3559dfb"
# if unitId==None:
#     print( "no unit id passed")
#     exit()

Url = 'https://data.exactspace.co/exactapi'
config={}
config['api']={}
config['api']['meta']=Url
config['api']['query']='https://data.exactspace.co/kairosapi/api/v1/datapoints/query'
config["api"]["datapoints"]='https://data.exactspace.co/exactdata/api/v1/datapoints'

def generateDateTime():
    format = "%Y-%m-%dT%H:%M:%S.%fZ"
    currentTime = datetime.datetime.now(pytz.timezone('Asia/Calcutta'))
    # Get the current time in the 'Asia/Calcutta' timezone

    # Subtract 8 hours from the current time
    time_diff = datetime.timedelta(hours=24)
    new_time = (currentTime - time_diff)
    
    return currentTime.strftime(format), new_time.strftime(format)

print("----")
def task_url():
    url = 'https://data.exactspace.co/exactapi/units/60ae9143e284d016d3559dfb/activities?filter={"order":"createdOn DESC","limit":100,"value":"Recommendation","fields":["status","createdOn","content","updateHistory","id"]}'
    # url="https://data.exactspace.co/exactapi/units/60ae9143e284d016d3559dfb/activities"
    response = requests.get(url)
    # if(response.status_code==200):
    tasks =json.loads(response.content)
    print(len(tasks))
    # print(tasks)
    return tasks

def shift_wise_task(startTime,endtime,tasks):
    from datetime import datetime, timedelta

    # startTime = '2024-08-01T00:00:00.000Z'
    # endtime =   '2024-08-01T12:00:00.000Z'
    start_date = datetime.strptime(startTime,"%Y-%m-%dT%H:%M:%S.%fZ")
    end_date = datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%fZ")
    # print(end_date)
    # Filter recommendations based on the date range
    filtered_tasks = [
        task for task in tasks
        if start_date <= datetime.strptime(task['createdOn'], '%Y-%m-%dT%H:%M:%S.%fZ') <= end_date]
    
    pprint.pprint(filtered_tasks)
    print(len(filtered_tasks))
    return filtered_tasks


def filter_Recommendation(filtered_tasks):
    data = []
    for task in filtered_tasks:
        if "status" in task:
            created_on = task['createdOn']
            title = next((item['value'] for item in task['content'] if item['type'] == 'title' and 'Recommendation' in item['value']), None)

            # Find all tables with the header ['Parameter', 'Actual Value', 'Recommended Value']
            parameter_tables = [item['value'] for item in task['content'] if item['type'] == 'table' and len(item['value']) > 0 and item['value'][0] == ['Parameter', 'Actual Value', 'Recommended Value']]

            update_history = task['updateHistory']
            commented_action = update_history[1]['action'] if len(update_history) > 1 else None
            completed_action = next((item['action'] for item in update_history if 'completed this task' in item['action']), None)

            status = task['status']

            # Process each parameter table
            for parameters in parameter_tables:
                if len(parameters) > 1:  # Ensure there are rows to process
                    for param in parameters[1:]:  # Skip header row
                        if len(param) >= 3:  # Ensure the row has enough columns
                            idx=task['id']
    #                         print(idx)
                            data.append({
                                'createdOn': created_on,
                                'Title': title,
                                'Parameter': param[0],
                                'Actual Value': param[1],
                                'Recommended Value': param[2],
                                'commented action': commented_action,
                                'Stakeholder': completed_action,
                                'status': status,
                                'Tasklink':(f'https://data.exactspace.co/pulse-master/my-tasks/{idx}')
                            })

    df = pd.DataFrame(data)

    if 'commented action' and 'Parameter'  in df:
        df['Parameter'] = df['Parameter'].str.replace('Mould 1', '', regex=False).str.replace('Mould 1', '', regex=False)

        df['commented action'] = df['commented action'].str.replace('<p>', '', regex=False).str.replace('</p>', '', regex=False)
        
        # print("Recommendation data not available in this shift")
    print(df,"filter_Recommendation_df")
    return df

def to_hh_mm(date_string):
    from datetime import datetime, timedelta

    date_obj = datetime.strptime(date_string, "%Y-%m-%dT%H:%M:%S.%fZ")
    return date_obj.strftime("%d-%m-%Y %H:%M")

def to_epoch_ist(date_string):
    from datetime import datetime, timedelta

    date_obj = datetime.strptime(date_string, "%d-%m-%Y %H:%M")
    # Adjust for IST (UTC+5:30)
    ist_offset = timedelta(hours=5, minutes=30)
    date_obj_ist = date_obj - ist_offset
    # Convert to epoch in milliseconds
    epoch_timestamp_ms = int(date_obj_ist.timestamp() * 1000)
    return epoch_timestamp_ms

def Recommendation_time(df):
    if "createdOn" in df:
        df['Recommendation time'] = df['createdOn'].apply(to_hh_mm)
        df['epoch'] = df['Recommendation time'].apply(to_epoch_ist)
    
def formatResultAsDF2(resultset):
    if(resultset):
        if(isinstance(resultset["queries"], list)):
            resultset["results"] = []
            f=0
            for res in resultset["queries"]:
                #print res["results"][0]
                if(len(res["results"][0]["values"])!=0):
                    df = pd.DataFrame(res["results"][0]["values"], columns=["time", res["results"][0]["name"]])
                    #print df.shape[0]
                    f=1
                    try:
                        if(final.shape[0] > df.shape[0]):
                            final = pd.merge_asof(final, df, on="time", tolerance=60000, direction="nearest")
                        else: 
                            final = pd.merge_asof(df, final, on="time", tolerance=60000, direction="nearest")
                    except Exception as e:
                        #print e
                        final = df
            if(f==1):
                resultset["results"].append({"data" : final})
                return resultset
            else: 
                resultset["results"].append({"data" : pd.DataFrame()})
                return resultset
        else:
            resultset["results"].append({"data" : pd.DataFrame()})
            return resultset
    else:
        return {}

def getdata_api2(taglist,start):

    url=config["api"]["query"]
    body ={
    "metrics": [],
    "plugins": [],
    "cache_time": 0,
    "start_absolute": start,
"end_absolute": start+ 3.6e+6
}
#        "start_relative": {
#   "value": "1",
#   "unit": "years"
# }
    # }

    for tag in taglist:
        query = {
            "tags": {},
            "name": tag,
            "aggregators": [{
                "name": "avg",
                "sampling": {
                    "value": "1",
                    "unit": "weeks"
                }
#                   {
#           "name": "gaps",
#           "sampling": {
#             "value": "1",
#             "unit": "minutes"
#           }              }]
        }]}
        body['metrics'].append(query)
    # print(body)

    res = requests.post(url=url,json=body)
    # print(res)
    if res.status_code == 200:
        resultset = json.loads(res.content)
    else:
        print ("query / url / post object something wrong")
    if resultset:
        resultset = formatResultAsDF2(resultset)
        if (len(resultset["results"]) > 0):
            df = resultset["results"][0]["data"]
            
    return df

def Geometric_Density(df):
    df['Geometric Density'] = None
    for index, row in df.iterrows():
        timestamp = row['epoch']
        result = getdata_api2(['GAP_GAP04.PLC04.MLD1_DATA_Anode_Geometric'], timestamp)
        try:
            first_value = round(result['GAP_GAP04.PLC04.MLD1_DATA_Anode_Geometric'].iloc[0],4) 
            # print(first_value)
            df.at[index, 'Geometric Density'] = first_value  
        except:
            pass
    df['Density Achieved after recommendations'] = df['Geometric Density'].apply(
        lambda x: 'Yes' if x >= 1.65 else 'No')
    # print(df,"Geometric_Density")
    return df

def getValues(startTime,endtime,taglist):
    from datetime import datetime, timedelta

    url = "https://data.exactspace.co/kairosapi/api/v1/datapoints/query"
    ist = pytz.timezone('Asia/Kolkata')
    # startTime = '2024-08-01T00:00:00.000Z'
    # endtime ='2024-08-01T12:00:00.000Z'
    # Convert the given date strings to datetime objects
    start_date = datetime.strptime(startTime,"%Y-%m-%dT%H:%M:%S.%fZ")
    end_date = datetime.strptime(endtime,"%Y-%m-%dT%H:%M:%S.%fZ")
    start_date_ist = ist.localize(start_date)
    end_date_ist = ist.localize(end_date)
    # Convert datetime objects to epoch time in milliseconds
    start_absolute = int(start_date_ist.timestamp() * 1000)
    end_absolute = int(end_date_ist.timestamp() * 1000)
    print(end_absolute,start_absolute,"$$$$$$$$$$$$$$$$$$$$$")
    d = {
        "metrics": [
            {
                "tags": {},
                "name": "",
            }
        ],
        "plugins": [],
        "cache_time": 0,
        "start_absolute": start_absolute,
        "end_absolute": end_absolute
    }
    
    finalDF = pd.DataFrame()
    for tag in taglist:
        d['metrics'][0]['name'] = tag
        res = requests.post(url=url, json=d)
        values = json.loads(res.content)
        df = pd.DataFrame(values["queries"][0]["results"][0]['values'], columns=['time', values["queries"][0]["results"][0]['name']])
        finalDF = pd.concat([finalDF, df], axis=1)

    finalDF = finalDF.loc[:, ~finalDF.columns.duplicated()]
    finalDF.dropna(subset=['time'], inplace=True)
    finalDF['time'] = pd.to_datetime(finalDF['time'], unit='ms')+ pd.Timedelta(hours=5.5)
    finalDF['time'] = finalDF['time'].dt.strftime('%d-%m-%y %H:%M')

    return finalDF

def actionTaken_value_to_30min(df,df2):
    if "Recommendation time" in df:
        df['Recommendation time'] = pd.to_datetime(df['Recommendation time'],dayfirst=True)
        # print(df['Recommendation time'])
        df2['time'] = pd.to_datetime(df2['time'],dayfirst=True)

        # Add createdOn_plus_30 column (just adding 30 minutes to the date will not make sense, use a fixed time or adjust accordingly)
        df['createdOn_plus_30'] = df['Recommendation time']  # Simplified for date only
        df['createdOn_plus_30'] = df['Recommendation time'] + pd.to_timedelta(30, unit='m')
        # Add new column 'Action Taken(current value)'
        df['Action Taken(current value)'] = None

        # Function to find the best matching column name
        def get_best_match(parameter, columns):
            match, score = process.extractOne(parameter, columns)
            return match if score > 80 else None  # Adjust threshold as needed

        # Function to find the value after 30 minutes
        def find_value_after_30min(row):
            parameter = row['Parameter']
            created_on_plus_30 = row['createdOn_plus_30']
        #     print(created_on_plus_30)
            best_match = get_best_match(parameter, df2.columns)
            if best_match is not None:
                df2_filtered = df2[df2['time'] >= created_on_plus_30]
                if not df2_filtered.empty:
                    closest_time_row = df2_filtered.iloc[0]
                    return closest_time_row[best_match]
            return None

        # Apply the function to each row in df
        df['Action Taken(current value)'] = round(df.apply(find_value_after_30min, axis=1),2)
        print(df,"###############")
    return df

def save_excel_file(new_df,new_column_order):
    df_processed = new_df.copy()
    # df_processed = new_df.groupby(new_column_order).first()
    # df_processed
    df_processed.loc[new_df.duplicated(subset=['Recommendation time', 'Title','Tasklink']), ['Recommendation time', 'Title','Tasklink']] = ""
    print(df_processed)
    # Write to Excel
    df_processed.to_excel("recommendations.xlsx", index=False)

    # Load the workbook and select the sheet
    file_path = 'recommendations.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Define the green fill for conditional formatting
    green_fill = PatternFill(start_color="C9DF8A", end_color="C9DF8A", fill_type="solid")

    # Define the green fill for the first row
    first_row_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Apply the green fill to the first row
    for cell in sheet[1]:
        cell.fill = first_row_fill
        cell.border = thin_border

    # Iterate over the rows and apply conditional formatting and border style
    for row in sheet.iter_rows(min_row=2, max_col=6, max_row=sheet.max_row):
        for cell in row:
            cell.border = thin_border

        recommended_value = row[4].value  # Recommended Value (E column)
        action_taken_value = row[5].value  # Action Taken (F column)

        # Extract numeric values from strings if necessary and compare
        try:
            # Check if "increase to" or "decrease to" is in the recommended value
            if 'increase to' in recommended_value.lower():
                recommended_value_numeric = float(recommended_value.split()[-1])
                action_taken_value_numeric = float(action_taken_value) if action_taken_value is not None else 0
                if action_taken_value_numeric >= recommended_value_numeric:
                    row[5].fill = green_fill
            elif 'decrease to' in recommended_value.lower():
                recommended_value_numeric = float(recommended_value.split()[-1])
                action_taken_value_numeric = float(action_taken_value) if action_taken_value is not None else 0
                if action_taken_value_numeric <= recommended_value_numeric:
                    row[5].fill = green_fill

        except (ValueError, IndexError, AttributeError):
            # Handle cases where conversion to float fails or values are not as expected
            pass

    # Format the Recommendation time column
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        recommendation_time_cell = row[0]  # Assuming Recommendation time is the first column (A)
        try:
            # Convert the cell value to a datetime object if it is not already
            if isinstance(recommendation_time_cell.value, datetime):
                dt = recommendation_time_cell.value
            else:
                dt = datetime.strptime(recommendation_time_cell.value, '%Y-%m-%d %H:%M:%S')
            
            # Format the datetime object to the desired format
            formatted_time = dt.strftime('%Y-%m-%d %H:%M:%S')
            recommendation_time_cell.value = formatted_time
        except (ValueError, TypeError):
            # Handle cases where the datetime parsing fails
            pass

    # Save the updated workbook
    output_file_path = 'Daily report-to-monitor-actions-taken-on-GAP-recommendations.xlsx'
    workbook.save(output_file_path)
    print(f"Updated file saved to {output_file_path}")
    return output_file_path

def uploadDataToAttachment(output_file_path):
    fileName = output_file_path
    path = "./"
    files = {'upload_file': open(str(path+fileName),'rb')}
    url =config["api"]["meta"]+ '/attachments/tasks/upload'
    # url= 'https://data.exactspace.co/exactapi' +'/attachments/tasks/upload'
    response = requests.post(url, files=files)
    status=""
    if(response.status_code==200):
        status="File uploaded to attachment"
        try:
            # os.remove(fileName)
            status+=" and also Removed from local directory"
        except:
            return "Uploaded to attachment but Something went wrong in removing file from local directory"
    else:
        time.sleep(10)
        response = requests.post(url, files=files)
        if(response.status_code==200):
            status="File uploaded to attachment"
            try:
                # os.remove(fileName)
                status+=" and also Removed from local directory"
            except:
                return "Uploaded to attachment but Something went wrong in removing file from local directory"
        else:
            return "Error! File upload Failed! File is in your local directory , error code: "+str(response.status_code)
    return status 

def timestamp_to_date(timestamp):
    return time.strftime('%Y-%m-%d', time.localtime(timestamp))


def send_mail(output_file_path):
    formatted_date = timestamp_to_date(int(time.time()))
    logopath=config["api"]["meta"]+'/attachments/mail/download/logo.png'
    html=' <!doctype html><html><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\"><title>Pulse </title><style type=\"text/css\">body{margin:0;}body,table,td,p,a,li,blockquote{-webkit-text-size-adjust:none!important;font-family:sans-serif;font-style:normal;font-weight:400;}button{width:90%;}@media screen and (max-width:600px){body,table,td,p,a,li,blockquote{-webkit-text-size-adjust:none!important;font-family:sans-serif;}table{width:100%;}.footer{height:auto!important;max-width:48%!important;width:48%!important;}.table.responsiveImage{height:auto!important;max-width:30%!important;width:30%!important;}.table.responsiveContent{height:auto!important;max-width:66%!important;width:66%!important;}.top{height:auto!important;max-width:48%!important;width:48%!important;}.catalog{margin-left:0%!important;}}@media screen and (max-width:480px){body,table,td,p,a,li,blockquote{-webkit-text-size-adjust:none!important;font-family:sans-serif;}table{width:100%!important;border-style:none!important;}.footer{height:auto!important;max-width:96%!important;width:96%!important;}.table.responsiveImage{height:auto!important;max-width:96%!important;width:96%!important;}.table.responsiveContent{height:auto!important;max-width:96%!important;width:96%!important;}.top{height:auto!important;max-width:100%!important;width:100%!important;}.catalog{margin-left:0%!important;}button{width:90%!important;}}</style></head><body yahoo=\"yahoo\" background=\"#f7f7f7\" style=\"background:#f7f7f7;\"><table width=\"100%\" cellspacing=\"0\" cellpadding=\"0\"><tbody><tr><td><table width=\"650\" align=\"center\" cellpadding=\"0\" cellspacing=\"0\" background=\"#fff\" style=\"background:#fff\"><tbody><tr><td bgcolor=\"#f7f7f7\"><table class=\"top\" width=\"48%\" align=\"center\" cellpadding=\"0\" cellspacing=\"0\" style=\"padding:40px 10px 10px 10px;\"><tbody><tr><td style=\"font-size:12px;color:#929292;text-align:center;font-family:sans-serif;padding-bottom:15px;\"><img src="'
    html+=logopath
    html+='" width="150"/></td></tr></tbody></table></td></tr><tr> <td style="border-bottom: solid 1px #CACACA; padding: 15px 15px 30px 15px"><table width="100%" align="left"  cellpadding="0" cellspacing="0"><tr><td width="250" style="padding-top:15px;"><div style="font-size:20px; padding:5px 0px;"><div id="incInfo"><span style="font-size:30px; padding:5px 0px;">'
    html+=str("Daily report-to-monitor-actions-taken-on-GAP-recommendations").replace("_"," ")
    html+='</span></div><div><small>Created on <b><i>'
    html+=formatted_date
    html+=' </i></b></small></div></td></tr></table></td></tr><tr> <td><table class="table" width="96%" align="center" style="text-align: center; padding-top: 15px;"><tbody><tr><td colspan="3" align="left" style="border-bottom: solid 1px #CACACA; padding-bottom: 8px; font-size: 15px;"><b> '
    html+="Customer Details"
    html+=' </b></td></tr><tr style="font-size: 13px; color: #9C9C9C"><td align="left" width="33.3%" style="padding-top: 10px;">Unit</td><td align="left" width="33.3%" style="padding-top: 10px;">Site</td><td align="left" width="33.3%" style="padding-top: 10px;">Customer</td></tr><tr><td align="left" width="33.3%"> '
    html+="GAP"
    html+='</td><td align="left" width="33.3%">'
    html+=str("GAP, Mahan")
    html+='</td><td align="left"  width="33.3%">'
    html+=str("GAP")
    html+='</td></tr></tbody></table><tr> <td style="padding:30px 0px 10px 0px; border-top:solid 1px #CACACA;"><table width="100%" align="left"  cellpadding="0" cellspacing="0"><tr><td align="center"><div> '
    mailrepname=str("Daily report-to-monitor-actions-taken-on-GAP-recommendations")

    file = output_file_path
    # time.sleep(2)
    
    f1='/src/uploads/tasks/'+output_file_path
    Report_name="Daily report-to-monitor-actions-taken-on-GAP-recommendations"
    emails=['dibyendu.g@adityabirla.com',
            'sayan.dey@adityabirla.com',
            'anurag.gaurav@adityabirla.com',
            'aswini.mishra@adityabirla.com']
    
    regards='ExactSpace Technologies</b></html>'
    body = {
        "to":emails,
        "subject":  str(mailrepname)+" "+str("GAP"),
        "html": html+'</div></td></tr></table></td></tr><tr> <td ><table width="96%" align="left" cellpadding="0" cellspacing="0"><tr><td style="border-bottom: solid 1px #CACACA; padding-bottom: 35px; padding-left: 15px; font-size: 20px;"> <b>Dear Sir/Ma\'am,<br>Find the attached '+Report_name.replace("_"," ")+' file in the mail.</b></td></tr></table></td></tr></tbody> </table></td></tr></tbody></table></body><br>Regards,<br><b>'+regards,
        "f1":f1,  
        "f2":"", 
        "f3":"", 
        "cc":['rahul.k@exactspace.co','ashlin.f@exactspace.co','nikhil.s@exactspace.co','kashmeen.a@exactspace.co'],
        "bcc":[] 
    }
    # print(body)
    # ,'nikhil.s@exactspace.co','ashlin.f@exactspace.co','arun@exactspace.co','sayan.dey@adityabirla.com'
    time.sleep(1)
    email = mg.Email()
    mailstatus=email.sendSESMailWithAttach(body)
    # print(mailstatus,"mailstatus")


########### calling all function in  main()################
def main():
    endtime, startTime = generateDateTime() 
    print("func called at ", endtime, startTime, endtime)
    
    tasks = task_url()
    filtered_tasks = shift_wise_task(startTime,endtime,tasks)
    df = filter_Recommendation(filtered_tasks)
    
    Recommendation_time(df)
    
    df = Geometric_Density(df)
    
    tags = [
        "GAP_GAP04.PLC04.K363_K010_LIC_01_LK_L2",
        "GAP_GAP04.PLC04.K363_K010_VKMIN",
        "GAP_GAP03.PLC03.ACTUAL_FORMULA.KBS",
        "GAP_GAP03.PLC03.ACTUAL_FORMULA.KLP",
        "GAP_GAP03.PLC03.ACTUAL_FORMULA.KGS",
        "GAP_GAP04.PLC04.K363_K030_WIT_01_WSP_OUT_PV",
        "GAP_GAP03.PLC03._362_J150_WIT_01.PV",
        "GAP_GAP04.PLC04.U363_K010_TT_01_PV",
        "GAP_GAP04.PLC04.MLD1_DATA_Anode_Weight",
        "GAP_GAP03.PLC03._362_J155_FT_01.PV",
        "GAP_GAP03.PLC03.U362_J155_FT_01_PV",
        "GAP_GAP04.PLC04.MLD2_DATA_Anode_Weight",
        "GAP_GAP01.PLC01._GAPPOS2.PV",
        "GAP_GAP01.PLC01._362_E020_MVF_01.ACTRL.AUTOSPEEDREF",
        "GAP_GAP04.PLC04.EU1_DATA_Anode_Counter_Pres",
        "GAP_GAP04.PLC04.EU2_DATA_Anode_Counter_Pres",
        "GAP_GAP04.PLC04.EU1_DATA_Anode_Vaccum_Pres",
        "GAP_GAP04.PLC04.EU2_DATA_Anode_Vaccum_Pres",
        "GAP_GAP04.PLC04.K363_K030_WIT_01_WIK_MLD1_NumUsed",
        "GAP_GAP04.PLC04.K363_K030_TWKD",
        "GAP_GAP04.PLC04.K010_WIT_01_PV"
    ]
    
    df2 = getValues(startTime,endtime,tags)
    df2.columns = [
        "time",
        "Feed hopper level setpoint",
        "Minimum high extraction rate set point",
        "Butt percent formula",
        "Pitch percentage",
        "Actual green scrap percentage",
        "Current paste weight setpoint",
        "mixer weight",
        "Paste feeder paste temperature",
        "Anode Weight",
        "Water flow rate",
        "Water flow outlet of cooler",
        "Mould 2 Anode Weight",
        "Rhodax gap",
        "Rhodax speed",
        "Anode Counter Pres",
        "Mould 2 Anode Counter Pres",
        "Anode Vaccum_Pres",
        "Mould 2 Anode Vaccum_Pres","Weighing transfer hopper","Transfer Hopper tare weight","Paste feeder weight"
    ]
    
    print(df2)
    df = actionTaken_value_to_30min(df, df2)
    
    new_column_order = [
        'Recommendation time',
        'Title',
        'Parameter',
        'Actual Value',
        'Recommended Value',
        'Action Taken(current value)',
        'commented action',
        'Stakeholder',
        'Density Achieved after recommendations',
        'Geometric Density',
        'status',"Tasklink"
    ]
    new_df = df.reindex(columns=new_column_order)
    # print(new_df.shape,"#############################################")
    output_file_path = save_excel_file(new_df, new_column_order)

    uploadDataToAttachment(output_file_path)
    send_mail(output_file_path)

main()
 
scheduler = BackgroundScheduler(timezone=pytz.timezone('Asia/Calcutta'))

# # scheduler.add_job(func=generateDateTime, trigger="interval", hours=6, args=[unitsId])
scheduler.add_job(main, trigger='cron', hour=18, minute=00, second=0)
# # scheduler.add_job(main, trigger='cron', hour=13, minute=30, second=0)
# # scheduler.add_job(main, trigger='cron', hour=13, minute=31, second=0)
scheduler.start()




##  TO see the logs on the console
# Using Event to keep the script running
stop_event = threading.Event()

try:
    stop_event.wait()  # Wait indefinitely until the event is set
except (KeyboardInterrupt, SystemExit):
    scheduler.shutdown()




